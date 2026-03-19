# ============================================================
# IMPORTS — doivent être en premier
# ============================================================
import re
import math
import io
import os
import glob
import hashlib
import json
from datetime import datetime

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
FILES_DIR       = os.path.join(BASE_DIR, "Files")
FILES_UPDATE_DIR = os.path.join(BASE_DIR, "Files to update")
SCRIPTS_DIR     = os.path.join(BASE_DIR, "Scripts")
ACP_FEED        = os.path.join(FILES_UPDATE_DIR, "ACP_OpenAI_Feed.csv")
ETIQUETTE_AB    = os.path.join(FILES_DIR, "Etiquette-AB - final.csv")
ETIQUETTE_CD    = os.path.join(FILES_DIR, "Etiquette-CD - final.csv")
RELATED_JSON    = os.path.join(SCRIPTS_DIR, "related_products.json")
GMC_PATTERN     = os.path.join(FILES_UPDATE_DIR, "Flux Google*")


def find_gmc_file():
    files = glob.glob(GMC_PATTERN)
    return files[0] if files else None


# ============================================================
# CHAMP MAP GMC — seules ces colonnes sont écrasées par l'ERP
# (mapping : colonne GMC → colonne ERP)
# ============================================================
GMC_FIELD_MAP = {
    # clé = colonne GMC (avec espaces)   valeur = colonne ERP (avec underscores)
    "title"                 : "title",
    "link"                  : "url",
    "image link"            : "image_link",
    "additional image link" : "additional_image_link",
    "item group id"         : "item_group_id",
    # description, availability, price, sale price → traitement spécial ci-dessous
}
GMC_SPECIAL = {"description", "availability", "price", "sale price"}
GMC_MANAGED = set(GMC_FIELD_MAP.keys()) | GMC_SPECIAL   # périmètre complet


# ============================================================
# HELPERS — fonctions utilitaires
# ============================================================

def is_empty(val):
    if val is None:
        return True
    if isinstance(val, float) and math.isnan(val):
        return True
    return str(val).strip() in ("", "nan", "None")


def clean_html(raw):
    if is_empty(raw):
        return ""
    text = re.sub(r"<[^>]+>", " ", str(raw))
    return re.sub(r"\s+", " ", text).strip()


def extract_bullets(html):
    if is_empty(html) or str(html).strip() == "<ul></ul>":
        return []
    items = re.findall(r"<li[^>]*>(.*?)</li>", str(html), re.IGNORECASE | re.DOTALL)
    return [clean_html(item) for item in items if clean_html(item)]


def build_description(description, rich_text):
    """
    Construit la description :
      <description ERP> Ses principaux bienfaits : <bienfait1>, <bienfait2>, ...
    - base = champ 'description' de l'ERP (texte court)
    - le point final de base est géré pour éviter la double ponctuation
    - bienfaits = <li> de rich_text_description, en minuscules, sans HTML
    """
    base = str(description).strip() if not is_empty(description) else ""
    base_clean = base.rstrip(".")           # BUG 3 corrigé
    bullets = extract_bullets(rich_text)
    if bullets:
        bienfaits = ", ".join(b.lower() for b in bullets)
        if base_clean:
            return f"{base_clean}. Ses principaux bienfaits : {bienfaits}"
        return f"Ses principaux bienfaits : {bienfaits}"
    return base


def normalize_availability(val):
    """AVAILABLE/in_stock → in_stock | OUT_OF_STOCK/AVAILABLE_SOON → out_of_stock"""
    if is_empty(val):
        return "out_of_stock"
    v = str(val).strip().lower()
    return "in_stock" if v in ("available", "in_stock") else "out_of_stock"


def format_price(val):
    """Retourne '12,90 EUR' (virgule décimale). BUG 2 corrigé."""
    if is_empty(val):
        return ""
    try:
        num = float(str(val).replace("EUR", "").replace(",", ".").strip())
        return f"{num:.2f}".replace(".", ",") + " EUR"
    except (ValueError, TypeError):
        return ""


def convert_links(val):
    """Liens séparés par ';' (ERP) → séparés par ', ' (GMC/ACP)."""
    if is_empty(val):
        return ""
    parts = [p.strip() for p in str(val).replace(";", ",").split(",") if p.strip()]
    return ", ".join(parts)


def extract_hyperlink_label(cell_value):
    """Extrait le label d'une formule =HYPERLINK('url','label')."""
    if cell_value and str(cell_value).startswith("=HYPERLINK"):
        m = re.search(r',\s*"([^"]+)"\s*\)', str(cell_value))
        if m:
            return m.group(1)
    return cell_value


def cell_val(v):
    if v is None:
        return ""
    if isinstance(v, float) and math.isnan(v):
        return ""
    return v


# ============================================================
# LECTURE DU GMC VIA OPENPYXL (préserve les HYPERLINK headers)
# BUG 6,7,8 corrigés
# ============================================================

def read_gmc_openpyxl(path, sheet_name="Feuille 1"):
    """Lit le fichier GMC et retourne (headers_labels, list_of_dicts)."""
    wb = load_workbook(path, data_only=False)
    # Chercher la feuille principale (Feuille 1 ou première feuille hors Produits_retirés)
    sheet = None
    for sn in wb.sheetnames:
        if sn not in ("Produits_retirés", "exad"):
            sheet = wb[sn]
            break
    if sheet is None:
        return [], []

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    raw_headers = rows[0]
    headers = [extract_hyperlink_label(h) or f"col_{i}" for i, h in enumerate(raw_headers)]

    data = []
    for row in rows[1:]:
        record = {h: cell_val(v) for h, v in zip(headers, row)}
        data.append(record)

    return headers, data


# ============================================================
# EXTRACTION CODES-BARRES
# ============================================================

def extract_barcodes():
    barcodes = {}
    for path in [ETIQUETTE_AB, ETIQUETTE_CD]:
        if not os.path.exists(path):
            continue
        et = pd.read_csv(path, dtype=str, keep_default_na=False)
        for _, row in et.iterrows():
            sku = row.get("sku", "").strip()
            url = row.get("codebarre", "").strip()
            if sku and url:
                match = re.search(r"/(\d{13})_", url)
                if match:
                    barcodes[sku] = match.group(1)
    return barcodes


# ============================================================
# HELPERS ACP
# ============================================================

def fix_price(val):
    """Normalise un prix au format '8,15 EUR'."""
    if not val or str(val).strip() == "":
        return ""
    v = str(val).strip()
    v = re.sub(r"\s*(EUR|€)\s*$", "", v, flags=re.IGNORECASE).strip()
    v = v.replace(".", ",")
    if re.match(r"^\d+,\d+$", v) or re.match(r"^\d+$", v):
        return v + " EUR"
    return v + " EUR"


def get_custom_variant_format(title):
    t = title.lower()
    if "poudre concentr" in t: return "Poudre concentree"
    if "gelules" in t or "gélules" in t: return "Gelules"
    if "grand sachet" in t: return "Grand Sachet"
    if "petit sachet" in t: return "Petit Sachet"
    if "sachets" in t or "sachet" in t: return "Sachet"
    if "flacon" in t: return "Flacon"
    return ""


def get_weight_from_title(title):
    match = re.search(r"(\d+)\s*g\b", title, re.IGNORECASE)
    return match.group(1) + "g" if match else ""


def build_variant_dict(title):
    variant = {}
    t = title.lower()
    if "poudre concentr" in t: variant["format"] = "Poudre concentree"
    elif "gelules" in t or "gélules" in t: variant["format"] = "Gelules"
    elif "grand sachet" in t: variant["format"] = "Grand Sachet"
    elif "petit sachet" in t: variant["format"] = "Petit Sachet"
    elif "sachets" in t or "sachet" in t: variant["format"] = "Sachet"
    w = re.search(r"(\d+\s*[gG](?:r)?)\b", title)
    if w: variant["poids"] = w.group(1).strip()
    return json.dumps(variant, ensure_ascii=False) if variant else ""


def build_group_title(title):
    parts = title.split(" - ")
    if len(parts) >= 2:
        return " - ".join(parts[:2]).strip()
    return title


# ============================================================
# Q&A & WARNING
# ============================================================

QA_LIST = [
    {"q": "Comment preparer et utiliser les plantes ?",
     "a": "Plantes entieres : decoction 20 min, boire apres les repas. Gelules : 6 par jour. Poudre concentree : diluer dans de l'eau chaude. Poudre moulue : faire bouillir 10 min."},
    {"q": "Combien de temps dure un traitement ?",
     "a": "Un traitement de 3 semaines est necessaire pour soulager les maux legers. Les premiers effets se ressentent generalement apres 3 semaines."},
    {"q": "Combien de temps dure une boite de 100 gelules ?",
     "a": "Environ 15 jours a raison de 6 gelules par jour."},
    {"q": "Peut-on utiliser les plantes pendant la grossesse ou l'allaitement ?",
     "a": "Verifiez la section Precautions d'emploi sur la fiche produit. En cas de doute, consultez votre praticien."},
    {"q": "Quelle est la quantite maximale par jour ?",
     "a": "En general, 13 g de poudre concentree ou 200 g de plantes entieres maximum par jour."},
    {"q": "Les gelules sont-elles d'origine naturelle ?",
     "a": "Oui, nos gelules sont en pullulane, une matiere naturelle sans risques pour la sante, sans transformation chimique et non allergene."},
    {"q": "D'ou proviennent les plantes ?",
     "a": "Toutes nos plantes proviennent de Chine ou elles sont cultivees avec le plus grand soin dans leurs regions de predilection, avec des controles qualite a l'export, l'import et en laboratoire."},
    {"q": "Combien de temps peut-on conserver les produits ?",
     "a": "Nos produits peuvent etre consommes plusieurs annees apres leur achat, a condition de les conserver a l'abri de l'humidite et du soleil."},
    {"q": "Les emballages sont-ils recyclables ?",
     "a": "Oui, les sachets en papier kraft et les flacons en plastique HDPE sont recyclables. Vous pouvez aussi rapporter les flacons vides en boutique ou par courrier."},
    {"q": "Quels sont les modes de livraison ?",
     "a": "Colissimo (48h), Chronopost (24h), points relais (48h), ou retrait en boutique au 15 rue de la Vistule, Paris 13. Livraison gratuite en point relais des 39 EUR."},
]

WARNING_STANDARD = (
    "Complement alimentaire deconseille aux enfants de moins de 12 ans. "
    "Deconseille aux femmes enceintes et allaitantes. "
    "Ne peut se substituer a une alimentation diversifiee et un mode de vie sain. "
    "Ne pas depasser la dose journaliere recommandee. "
    "Conserver au sec, a l'abri de la lumiere et de l'humidite. "
    "Tenir hors de portee des enfants."
)

ACP_COLUMNS = [
    "is_eligible_search", "is_eligible_checkout",
    "item_id", "gtin", "mpn", "identifier_exists", "title", "description", "url",
    "brand", "condition", "product_category", "material",
    "dimensions", "length", "width", "height", "dimensions_unit",
    "weight", "item_weight_unit", "age_group",
    "image_url", "additional_image_urls", "video_url", "model_3d_url",
    "price", "sale_price", "sale_price_start_date", "sale_price_end_date",
    "unit_pricing_measure", "pricing_trend",
    "availability", "availability_date", "expiration_date", "pickup_method", "pickup_sla",
    "group_id", "listing_has_variations", "variant_dict", "item_group_title",
    "color", "size", "size_system", "gender", "offer_id",
    "custom_variant1_category", "custom_variant1_option",
    "custom_variant2_category", "custom_variant2_option",
    "custom_variant3_category", "custom_variant3_option",
    "shipping", "is_digital",
    "seller_name", "marketplace_seller", "seller_url", "seller_privacy_policy", "seller_tos",
    "accepts_returns", "return_deadline_in_days", "accepts_exchanges", "return_policy",
    "popularity_score", "return_rate",
    "warning", "age_restriction",
    "review_count", "star_rating", "store_review_count", "store_star_rating",
    "q_and_a", "reviews",
    "related_product_id", "relationship_type",
    "target_countries", "store_country", "geo_price", "geo_availability",
]

# Mapping ERP → ACP
# Seules ces colonnes sont écrasées — les 68 autres colonnes ACP sont préservées
ACP_FIELD_MAP = {
    # clé = colonne ACP        valeur = colonne ERP
    "item_id"               : "id",            # clé de jointure (ne pas écraser, sert d'index)
    "title"                 : "title",
    "url"                   : "url",
    "image_url"             : "image_link",
    "additional_image_urls" : "additional_image_link",
    "group_id"              : "item_group_id",
    # description, availability, price, sale_price → traitement spécial ci-dessous
}
ACP_SPECIAL  = {"description", "availability", "price", "sale_price"}
ACP_MANAGED  = set(ACP_FIELD_MAP.keys()) | ACP_SPECIAL



# ============================================================
# HELPERS ACP — miroir des helpers GMC
# ============================================================

def update_acp_record(record, src):
    """
    Met à jour un enregistrement ACP depuis l'ERP.
    Respecte strictement ACP_MANAGED — les 68 autres colonnes ne sont PAS touchées.
    """
    updated = record.copy()

    # Champs directs (sauf item_id qui est la clé, ne pas écraser)
    for acp_col, erp_col in ACP_FIELD_MAP.items():
        if acp_col == "item_id":
            continue
        val = src.get(erp_col, None)
        if acp_col == "additional_image_urls":
            updated[acp_col] = convert_links(val)
        else:
            updated[acp_col] = "" if is_empty(val) else str(val).strip()

    # Champs spéciaux
    updated["description"]  = build_description(
        src.get("description", ""), src.get("rich_text_description", "")
    )
    updated["availability"] = normalize_availability(src.get("availability", ""))
    updated["price"]        = format_price(src.get("price", ""))
    sale_raw = src.get("sale_price", "")
    updated["sale_price"]   = format_price(sale_raw) if not is_empty(sale_raw) else ""

    return updated


def load_acp_as_df(path):
    """Lit le fichier ACP CSV et retourne un DataFrame (colonnes propres)."""
    if not os.path.exists(path):
        return pd.DataFrame()
    return pd.read_csv(path, dtype=str, keep_default_na=False)


def save_acp_from_df(df, path):
    """Sauvegarde le DataFrame ACP en CSV (préserve toutes les colonnes)."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8")


# ============================================================
# GÉNÉRATION ACP
# ============================================================

def generate_acp(xlsx_bytes, log):
    dx = pd.read_excel(io.BytesIO(xlsx_bytes), dtype=str, keep_default_na=False)
    dx["id"] = dx["id"].str.strip()
    erp = dx.set_index("id")
    erp_ids = set(erp.index)

    acp_old = load_acp_as_df(ACP_FEED)

    if not acp_old.empty and "item_id" in acp_old.columns:
        acp_old = acp_old.set_index("item_id")
        acp_ids = set(acp_old.index)

        # Produits retirés (dans ACP mais plus dans ERP)
        removed_ids = acp_ids - erp_ids
        removed_df  = acp_old.loc[list(removed_ids)].reset_index() if removed_ids else pd.DataFrame()

        # Mise à jour des produits existants
        for pid in acp_ids & erp_ids:
            updated = update_acp_record(acp_old.loc[pid].to_dict(), erp.loc[pid])
            for col, val in updated.items():
                if col in acp_old.columns:
                    acp_old.at[pid, col] = val

        # Nouveaux produits
        for pid in erp_ids - acp_ids:
            new_record = {col: "" for col in acp_old.columns}
            new_record["item_id"] = pid
            new_record = update_acp_record(new_record, erp.loc[pid])
            acp_old.loc[pid] = {col: new_record.get(col, "") for col in acp_old.columns}

        df = acp_old.reset_index().rename(columns={"index": "item_id"})
        if "item_id" not in df.columns:
            df = df.reset_index()

    else:
        # Première création
        rows = []
        for pid, src_row in erp.iterrows():
            record = {col: "" for col in ACP_COLUMNS}
            record["item_id"] = pid
            record = update_acp_record(record, src_row)
            rows.append(record)
        df = pd.DataFrame(rows, columns=ACP_COLUMNS)
        removed_df = pd.DataFrame()

    stats = {
        "total"    : len(df),
        "nouveaux" : len(erp_ids - (set(acp_old.index) if not acp_old.empty else set())),
        "retires"  : len(removed_df),
    }
    log.append(f"ACP: {len(df)} produits ({stats['nouveaux']} nouveaux, {stats['retires']} retirés)")
    return df, stats, removed_df


# ============================================================
# GÉNÉRATION GMC — BUG 5,6,7,8,12,13,14,15 corrigés
# ============================================================


def load_gmc_as_df(path):
    """Lit le fichier GMC via openpyxl (préserve les labels HYPERLINK) et retourne un DataFrame."""
    headers, data = read_gmc_openpyxl(path)
    if not data:
        return pd.DataFrame()
    return pd.DataFrame(data, columns=headers)


def save_gmc_from_df(df, path):
    """Réécrit le fichier GMC en préservant les en-têtes HYPERLINK d'origine."""
    wb = load_workbook(path)
    main_sheet = next((sn for sn in wb.sheetnames if sn not in ("Produits_retirés","exad")), wb.sheetnames[0])
    ws = wb[main_sheet]

    # Supprimer les données (ligne 2+), garder ligne 1 (HYPERLINK headers)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    normal_font = Font(name="Arial", size=10)
    # Récupérer l'ordre des colonnes depuis les headers actuels du fichier
    headers = [extract_hyperlink_label(ws.cell(1,c).value) for c in range(1, ws.max_column+1)]

    for _, row in df.iterrows():
        ws.append([row.get(h, "") for h in headers])

    for ws_row in ws.iter_rows(min_row=2):
        for cell in ws_row:
            cell.font = normal_font

    wb.save(path)

def update_gmc_record(record, src):
    """
    Met à jour un enregistrement GMC depuis l'ERP.
    Respecte strictement GMC_MANAGED — les autres colonnes ne sont PAS touchées.
    """
    updated = record.copy()

    # Champs directs
    for gmc_col, erp_col in GMC_FIELD_MAP.items():
        val = src.get(erp_col, None)
        if gmc_col == "additional image link":
            updated[gmc_col] = convert_links(val)     # BUG 13 corrigé
        else:
            updated[gmc_col] = "" if is_empty(val) else str(val).strip()

    # Champs spéciaux
    updated["description"] = build_description(
        src.get("description", ""), src.get("rich_text_description", "")
    )
    updated["availability"] = normalize_availability(src.get("availability", ""))  # BUG 15 corrigé
    updated["price"] = format_price(src.get("price", ""))                           # BUG 12 corrigé (une seule fois)
    sale_raw = src.get("sale_price", "")   # ERP utilise underscore
    updated["sale price"] = format_price(sale_raw) if not is_empty(sale_raw) else ""

    return updated


def generate_gmc(xlsx_bytes, log):
    # Charger ERP
    dx = pd.read_excel(io.BytesIO(xlsx_bytes), dtype=str, keep_default_na=False)
    dx["id"] = dx["id"].str.strip()
    erp = dx.set_index("id")
    erp_ids = set(erp.index)

    gmc_path = find_gmc_file()

    if gmc_path and os.path.exists(gmc_path):
        # Lire via openpyxl pour préserver les HYPERLINK headers (BUG 6,7,8 corrigés)
        headers, gmc_data = read_gmc_openpyxl(gmc_path)
        gmc_ids = {str(r.get("id","")).strip() for r in gmc_data}
        gmc_ids.discard("")

        removed_ids = gmc_ids - erp_ids
        active_rows = []
        removed_rows = []

        for record in gmc_data:
            pid = str(record.get("id","")).strip()
            if not pid:
                continue
            if pid in removed_ids:
                removed_rows.append(record)
                continue
            if pid in erp_ids:
                record = update_gmc_record(record, erp.loc[pid])
            active_rows.append(record)

        # Nouveaux produits (dans ERP mais pas dans GMC)
        new_ids = erp_ids - gmc_ids
        for pid in new_ids:
            src = erp.loc[pid]
            new_record = {h: "" for h in headers}
            new_record["id"] = pid
            new_record = update_gmc_record(new_record, src)
            active_rows.append(new_record)

        removed_df = pd.DataFrame(removed_rows, columns=headers) if removed_rows else pd.DataFrame()

        stats = {
            "total": len(active_rows),
            "nouveaux": len(new_ids),
            "retires": len(removed_rows),
        }
        log.append(f"GMC: {len(active_rows)} produits ({len(new_ids)} nouveaux, {len(removed_rows)} retirés)")

        # Écrire le résultat en préservant les HYPERLINK headers (BUG 5 corrigé)
        import shutil
        shutil.copy2(gmc_path, gmc_path + ".bak")
        wb = load_workbook(gmc_path)
        # Trouver la feuille principale
        main_sheet_name = next((sn for sn in wb.sheetnames if sn not in ("Produits_retirés","exad")), wb.sheetnames[0])
        ws = wb[main_sheet_name]

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        normal_font = Font(name="Arial", size=10)

        def make_row(record, hdrs):
            return [record.get(h, "") for h in hdrs]

        for record in active_rows:
            ws.append(make_row(record, headers))

        for ws_row in ws.iter_rows(min_row=2):
            for cell in ws_row:
                cell.font = normal_font

        # Feuille Produits_retirés
        removed_sheet = "Produits_retirés"
        if removed_sheet in wb.sheetnames:
            ws_r = wb[removed_sheet]
            existing_ids = {str(ws_r.cell(r, 1).value or "").strip() for r in range(2, ws_r.max_row+1)}
            for record in removed_rows:
                if str(record.get("id","")).strip() not in existing_ids:
                    ws_r.append(make_row(record, headers))
        else:
            ws_r = wb.create_sheet(removed_sheet)
            ws_r.append(headers)
            for cell in ws_r[1]:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = PatternFill("solid", start_color="FFD700")
            for record in removed_rows:
                ws_r.append(make_row(record, headers))

        for ws_row in ws_r.iter_rows(min_row=2):
            for cell in ws_row:
                cell.font = normal_font

        wb.save(gmc_path)
        log.append(f"GMC sauvegardé : {gmc_path}")

        # Retourner un DataFrame pour l'affichage
        gmc_df = pd.DataFrame(active_rows, columns=headers) if active_rows else pd.DataFrame()
        return gmc_df, stats, removed_df

    else:
        log.append("Aucun fichier GMC trouvé dans 'Files to update'")
        return None, {"total": 0, "nouveaux": 0, "retires": 0}, pd.DataFrame()


# ============================================================
# HELPERS UI
# ============================================================

def html_bar_chart(labels, values, color="#89B832"):
    if not values:
        return
    max_val = max(values) if max(values) > 0 else 1
    bars_html = ""
    for label, val in zip(labels, values):
        pct = val / max_val * 100
        bars_html += f"""
        <div style="display:flex;align-items:center;margin:6px 0;font-family:'Manrope',sans-serif;">
            <div style="width:180px;font-size:13px;color:#1A1A1A;text-align:right;padding-right:12px;flex-shrink:0;">{label}</div>
            <div style="flex:1;background:#E5E5E5;border-radius:4px;height:26px;overflow:hidden;">
                <div style="width:{pct}%;background:{color};height:100%;border-radius:4px;min-width:2px;"></div>
            </div>
            <div style="width:60px;font-size:13px;font-weight:600;color:#1A1A1A;padding-left:10px;">{val}</div>
        </div>"""
    st.markdown(f'<div style="margin:8px 0;">{bars_html}</div>', unsafe_allow_html=True)


def html_histogram(prices_series, bins=15, color="#89B832"):
    if prices_series.empty:
        return
    cuts = pd.cut(prices_series, bins=bins)
    hist = cuts.value_counts().sort_index()
    labels = [f"{iv.left:.0f}-{iv.right:.0f}" for iv in hist.index]
    html_bar_chart(labels, hist.values.tolist(), color)


def column_fill_stats(df):
    rows = []
    for col in df.columns:
        non_empty = (df[col] != "").sum()
        rows.append({
            "Colonne": col,
            "Remplis": f"{non_empty}/{len(df)}",
            "Taux": f"{non_empty/len(df)*100:.0f}%",
            "Vide": "Non" if non_empty > 0 else "Oui",
        })
    return pd.DataFrame(rows)


def availability_stats(df, col="availability"):
    if col not in df.columns:
        return pd.DataFrame()
    counts = df[col].value_counts().reset_index()
    counts.columns = ["Statut", "Nombre"]
    return counts


def price_stats(df, col="price"):
    if col not in df.columns:
        return {}
    raw = (df[col]
           .str.replace(r"\s*EUR\s*$", "", regex=True)
           .str.replace(",", ".")
           .str.replace(r"[^\d.]", "", regex=True))
    prices = pd.to_numeric(raw, errors="coerce").dropna()
    if prices.empty:
        return {}
    return {
        "Nb produits avec prix": int(prices.count()),
        "Prix min": f"{prices.min():.2f}",
        "Prix max": f"{prices.max():.2f}",
        "Prix moyen": f"{prices.mean():.2f}",
        "Prix median": f"{prices.median():.2f}",
    }


# ============================================================
# STREAMLIT — Configuration
# ============================================================

st.set_page_config(
    page_title="Calebasse - Mise a jour des feeds",
    page_icon="🌿",
    layout="wide",
)

# ============================================================
# AUTHENTIFICATION
# ============================================================
USERS = {
    "admin":      hashlib.sha256("Calebasse2026!".encode()).hexdigest(),
    "calebasse":  hashlib.sha256("feeds@Lab".encode()).hexdigest(),
}


def check_login():
    if st.session_state.get("authenticated"):
        return True
    st.markdown("""
    <style>[data-testid="stSidebar"] { display: none !important; }</style>
    """, unsafe_allow_html=True)
    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        st.markdown("""
        <div style="text-align:center;margin-top:60px;margin-bottom:30px;">
            <h1 style="font-family:Georgia,serif;color:#89B832;font-size:36px;margin:0;">Laboratoire Calebasse</h1>
            <p style="color:#666;font-style:italic;font-size:14px;margin:6px 0 0;">Gestion des feeds produits</p>
        </div>
        """, unsafe_allow_html=True)
        with st.form("login_form"):
            st.markdown('<p style="font-size:15px;font-weight:600;color:#1A1A1A;margin-bottom:4px;">Connexion</p>', unsafe_allow_html=True)
            username = st.text_input("Identifiant")
            password = st.text_input("Mot de passe", type="password")
            submitted = st.form_submit_button("Se connecter", type="primary", use_container_width=True)
            if submitted:
                pwd_hash = hashlib.sha256(password.encode()).hexdigest()
                if username in USERS and USERS[username] == pwd_hash:
                    st.session_state["authenticated"] = True
                    st.session_state["username"] = username
                    st.rerun()
                else:
                    st.error("Identifiant ou mot de passe incorrect")
        st.markdown('<div style="text-align:center;margin-top:40px;font-size:12px;color:#bbb;">Laboratoire Calebasse &copy; 2026</div>', unsafe_allow_html=True)
    return False


if not check_login():
    st.stop()

# ============================================================
# CSS
# ============================================================
st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Manrope:wght@400;600;700&display=swap" rel="stylesheet">
<style>
    :root, [data-testid="stAppViewContainer"], [data-testid="stHeader"],
    [data-testid="stSidebar"], .main, .block-container,
    [data-testid="stAppViewBlockContainer"] {
        background-color: #F9F9F7 !important;
        color: #1A1A1A !important;
    }
    [data-testid="stSidebar"] { background-color: #FFFFFF !important; }
    section[data-testid="stSidebar"] .stMarkdown,
    .main .stMarkdown, .main p, .main span, .main label, .main li,
    .main h1, .main h2, .main h3, .main h4 { color: #1A1A1A !important; }
    [data-testid="stExpander"] { background-color: #FFFFFF !important; border-color: #E5E5E5 !important; }
    [data-testid="stExpander"] summary span, [data-testid="stExpander"] p { color: #1A1A1A !important; }
    .stDataFrame, [data-testid="stTable"] { background-color: #FFFFFF !important; }
    [data-testid="stFileUploader"] { background-color: #FFFFFF !important; border-color: #E5E5E5 !important; }
    [data-testid="stFileUploader"] label, [data-testid="stFileUploader"] span,
    [data-testid="stFileUploader"] p { color: #1A1A1A !important; }
    .stTabs [data-baseweb="tab-list"] { background-color: transparent !important; }
    .stTabs [data-baseweb="tab"] { color: #1A1A1A !important; }
    .stTabs [aria-selected="true"] { color: #89B832 !important; border-bottom-color: #89B832 !important; }
    .stButton > button[kind="primary"],
    .stButton > button[data-testid="stBaseButton-primary"] {
        background-color: #89B832 !important; border-color: #89B832 !important; color: white !important;
    }
    .stButton > button[kind="primary"]:hover { background-color: #7aa52b !important; border-color: #7aa52b !important; }
    .stDownloadButton > button { border-color: #89B832 !important; color: #89B832 !important; }
    .stDownloadButton > button:hover { background-color: #f3fbe8 !important; }
    .calebasse-header { border-bottom: 3px solid #89B832; padding-bottom: 16px; margin-bottom: 28px; }
    .calebasse-header h1 { font-family: Georgia, serif; color: #89B832; margin: 0; font-size: 34px; }
    .calebasse-header p { margin: 4px 0 0; color: #666; font-style: italic; font-size: 15px; }
    .section-title { font-family: Georgia, serif; border-bottom: 1px solid #ddd; padding-bottom: 8px; color: #1A1A1A; font-size: 22px; margin-top: 28px; }
    [data-testid="stMetricValue"] { color: #89B832 !important; font-weight: 700 !important; }
    .success-box { background: #f3fbe8; border: 1px solid #89B832; border-radius: 6px; padding: 14px 18px; margin: 16px 0; font-weight: 600; color: #3d5a0f; }
    .calebasse-footer { margin-top: 50px; text-align: center; font-size: 13px; color: #999; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 10px 0 20px;">
        <h2 style="font-family: Georgia, serif; color: #89B832; margin: 0;">Calebasse</h2>
        <p style="color: #666; font-size: 13px; margin: 4px 0 0;">Gestion des feeds produits</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    page = st.radio(
        "Navigation",
        ["Importer le fichier Export ERP", "Apercu des fichiers actuels"],
        label_visibility="collapsed",
    )
    st.markdown("---")
    acp_exists = os.path.exists(ACP_FEED)
    gmc_exists = find_gmc_file() is not None
    st.markdown("**Fichiers detectes**")
    st.markdown(f"{'&#9989;' if acp_exists else '&#10060;'} ACP Feed", unsafe_allow_html=True)
    st.markdown(f"{'&#9989;' if gmc_exists else '&#10060;'} GMC Feed", unsafe_allow_html=True)
    st.markdown("---")
    st.caption(f"Connecte : **{st.session_state.get('username', '')}**")
    if st.button("Deconnexion", key="logout"):
        st.session_state["authenticated"] = False
        st.session_state["username"] = ""
        st.rerun()

st.markdown('<div class="calebasse-header"><h1>Calebasse - Mise à jour des feeds</h1><p>Outil de gestion des feeds ACP & GMC</p></div>', unsafe_allow_html=True)


# ============================================================
# PAGE 1 : Importer le fichier Export ERP
# ============================================================
if page == "Importer le fichier Export ERP":

    st.markdown('<h2 class="section-title">Importer le fichier Export ERP</h2>', unsafe_allow_html=True)
    uploaded = st.file_uploader(
        "Glissez ou selectionnez un fichier export-variants (.xlsx)",
        type=["xlsx"],
        help="Fichier exporte depuis l'ERP (ex: export-variants-2026-03-12.xlsx)",
    )

    if uploaded:
        xlsx_bytes = uploaded.read()

        preview_df = pd.read_excel(io.BytesIO(xlsx_bytes), dtype=str, nrows=5)
        total_rows = len(pd.read_excel(io.BytesIO(xlsx_bytes), dtype=str))
        with st.expander(f"Apercu du fichier importe ({total_rows} lignes, {len(preview_df.columns)} colonnes)", expanded=False):
            st.dataframe(preview_df, use_container_width=True)

        st.markdown("---")

        if st.button("Mettre a jour les feeds", type="primary", use_container_width=True):
            acp_log, gmc_log = [], []

            with st.spinner("Generation du feed ACP OpenAI..."):
                acp_df, acp_stats, acp_removed = generate_acp(xlsx_bytes, acp_log)

            with st.spinner("Generation du feed Google Merchant Center..."):
                gmc_df, gmc_stats, gmc_removed = generate_gmc(xlsx_bytes, gmc_log)

            # Sauvegarde ACP
            save_acp_from_df(acp_df, ACP_FEED)
            acp_log.append(f"Sauvegarde ACP : {ACP_FEED}")

            # La sauvegarde GMC est déjà faite dans generate_gmc() via openpyxl

            st.session_state.update({
                "acp_df": acp_df,
                "gmc_df": gmc_df,
                "acp_stats": acp_stats,
                "gmc_stats": gmc_stats,
                "acp_log": acp_log,
                "gmc_log": gmc_log,
                "acp_removed": acp_removed,
                "gmc_removed": gmc_removed,
                "generated": True,
            })

        if st.session_state.get("generated"):
            acp_df      = st.session_state.get("acp_df", pd.DataFrame())
            gmc_df      = st.session_state.get("gmc_df", pd.DataFrame())
            acp_stats   = st.session_state.get("acp_stats", {})
            gmc_stats   = st.session_state.get("gmc_stats", {})
            acp_log     = st.session_state.get("acp_log", [])
            gmc_log     = st.session_state.get("gmc_log", [])
            acp_removed = st.session_state.get("acp_removed", pd.DataFrame())
            gmc_removed = st.session_state.get("gmc_removed", pd.DataFrame())

            st.markdown('<div class="success-box">Mise a jour terminee — fichiers sauvegardes automatiquement</div>', unsafe_allow_html=True)

            # Stats
            st.markdown('<h2 class="section-title">Resultats</h2>', unsafe_allow_html=True)
            col1, col2 = st.columns(2, gap="large")

            with col1:
                st.markdown('<div style="background:#fff;border:1px solid #E5E5E5;border-left:4px solid #89B832;border-radius:8px;padding:18px 16px 10px;"><h3 style="margin:0 0 12px;font-size:16px;color:#89B832;font-family:Georgia,serif;">Feed ACP OpenAI</h3></div>', unsafe_allow_html=True)
                c1, c2, c3 = st.columns(3)
                c1.metric("Total", acp_stats.get("total", 0))
                c2.metric("Nouveaux", f"+{acp_stats.get('nouveaux', 0)}")
                c3.metric("Retires", f"-{acp_stats.get('retires', 0)}")

            with col2:
                st.markdown('<div style="background:#fff;border:1px solid #E5E5E5;border-left:4px solid #1A1A1A;border-radius:8px;padding:18px 16px 10px;"><h3 style="margin:0 0 12px;font-size:16px;color:#1A1A1A;font-family:Georgia,serif;">Feed Google Merchant Center</h3></div>', unsafe_allow_html=True)
                if gmc_df is not None:
                    g1, g2, g3 = st.columns(3)
                    g1.metric("Total", gmc_stats.get("total", 0))
                    g2.metric("Nouveaux", f"+{gmc_stats.get('nouveaux', 0)}")
                    g3.metric("Retires", f"-{gmc_stats.get('retires', 0)}")

            with st.expander("Journal des operations", expanded=False):
                st.markdown("**ACP:**")
                for line in acp_log:
                    st.text(line)
                st.markdown("**GMC:**")
                for line in gmc_log:
                    st.text(line)

            st.markdown("---")
            st.markdown('<h2 class="section-title">Apercu des donnees generees</h2>', unsafe_allow_html=True)

            tab_names = ["ACP OpenAI Feed", "Google Merchant Center"]
            if not gmc_removed.empty:
                tab_names.append(f"Produits retires ({len(gmc_removed)})")
            tabs = st.tabs(tab_names)

            # ── Onglet ACP (BUG 11 corrigé) ──────────────────────────────
            with tabs[0]:
                if not acp_df.empty:
                    st.markdown(f"**{len(acp_df)} produits, {len(acp_df.columns)} colonnes**")
                    acp_sub1, acp_sub2, acp_sub3, acp_sub4 = st.tabs([
                        "Tableau", "Colonnes & Remplissage", "Disponibilite", "Prix"
                    ])
                    with acp_sub1:
                        all_acp_cols = list(acp_df.columns)
                        acp_sel = st.multiselect("Colonnes a afficher", all_acp_cols, default=all_acp_cols, key="gen_acp_cols")
                        search_acp = st.text_input("Rechercher", key="gen_acp_search")
                        d = acp_df[acp_sel] if acp_sel else acp_df
                        if search_acp:
                            d = d[d.apply(lambda r: r.astype(str).str.contains(search_acp, case=False).any(), axis=1)]
                        st.dataframe(d, use_container_width=True, height=450)
                        st.caption(f"{len(d)} lignes")
                    with acp_sub2:
                        fill = column_fill_stats(acp_df)
                        m1, m2 = st.columns(2)
                        m1.metric("Colonnes remplies", f"{(fill['Vide']=='Non').sum()}/{len(acp_df.columns)}")
                        m2.metric("Colonnes vides", f"{(fill['Vide']=='Oui').sum()}/{len(acp_df.columns)}")
                        st.dataframe(fill, use_container_width=True, height=450)
                    with acp_sub3:
                        avail = availability_stats(acp_df)
                        if not avail.empty:
                            html_bar_chart(avail["Statut"].tolist(), avail["Nombre"].tolist())
                            st.dataframe(avail, use_container_width=True)
                    with acp_sub4:
                        ps = price_stats(acp_df, "price")
                        if ps:
                            pcols = st.columns(len(ps))
                            for i, (k, v) in enumerate(ps.items()):
                                pcols[i].metric(k, v)
                else:
                    st.info("Aucune donnee ACP generee")

            # ── Onglet GMC ────────────────────────────────────────────────
            with tabs[1]:
                if gmc_df is not None and not gmc_df.empty:
                    st.markdown(f"**{len(gmc_df)} produits, {len(gmc_df.columns)} colonnes**")
                    gmc_sub1, gmc_sub2, gmc_sub3, gmc_sub4 = st.tabs([
                        "Tableau", "Colonnes & Remplissage", "Disponibilite", "Prix"
                    ])
                    with gmc_sub1:
                        all_gmc_cols = list(gmc_df.columns)
                        gmc_sel = st.multiselect("Colonnes a afficher", all_gmc_cols, default=all_gmc_cols, key="gen_gmc_cols")
                        search_gmc = st.text_input("Rechercher", key="gen_gmc_search")
                        dg = gmc_df[gmc_sel] if gmc_sel else gmc_df
                        if search_gmc:
                            dg = dg[dg.apply(lambda r: r.astype(str).str.contains(search_gmc, case=False).any(), axis=1)]
                        st.dataframe(dg, use_container_width=True, height=450)
                        st.caption(f"{len(dg)} lignes")
                    with gmc_sub2:
                        gfill = column_fill_stats(gmc_df)
                        m1, m2 = st.columns(2)
                        m1.metric("Colonnes remplies", f"{(gfill['Vide']=='Non').sum()}/{len(gmc_df.columns)}")
                        m2.metric("Colonnes vides", f"{(gfill['Vide']=='Oui').sum()}/{len(gmc_df.columns)}")
                        st.dataframe(gfill, use_container_width=True, height=450)
                    with gmc_sub3:
                        gavail = availability_stats(gmc_df)
                        if not gavail.empty:
                            html_bar_chart(gavail["Statut"].tolist(), gavail["Nombre"].tolist())
                            st.dataframe(gavail, use_container_width=True)
                    with gmc_sub4:
                        gps = price_stats(gmc_df, "price")
                        if gps:
                            gpcols = st.columns(len(gps))
                            for i, (k, v) in enumerate(gps.items()):
                                gpcols[i].metric(k, v)
                            prices_g = pd.to_numeric(
                                gmc_df["price"].str.replace(r"\s*EUR\s*$","",regex=True).str.replace(",",".").str.replace(r"[^\d.]","",regex=True),
                                errors="coerce"
                            ).dropna()
                            html_histogram(prices_g)
                else:
                    st.warning("Fichier GMC non genere ou introuvable")

            # ── Onglet Produits retirés ───────────────────────────────────
            if not gmc_removed.empty and len(tabs) > 2:
                with tabs[2]:
                    st.markdown(f"**{len(gmc_removed)} produits retires**")
                    default_ret = ["id", "title", "price", "availability", "gtin", "brand"]
                    avail_ret = [c for c in default_ret if c in gmc_removed.columns]
                    st.dataframe(gmc_removed[avail_ret] if avail_ret else gmc_removed, use_container_width=True, height=450)

            st.markdown("---")
            st.markdown('<h2 class="section-title">Telecharger</h2>', unsafe_allow_html=True)
            dl1, dl2 = st.columns(2)

            with dl1:
                # Lire le fichier sauvegardé directement (source de vérité)
                if os.path.exists(ACP_FEED):
                    with open(ACP_FEED, "rb") as _fa:
                        acp_raw = _fa.read()
                else:
                    acp_raw = acp_df.to_csv(index=False, encoding="utf-8").encode("utf-8")
                st.download_button(
                    label="ACP_OpenAI_Feed.csv",
                    data=acp_raw,
                    file_name="ACP_OpenAI_Feed.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

            with dl2:
                if gmc_df is not None and not gmc_df.empty:
                    gmc_buffer = io.BytesIO()
                    with pd.ExcelWriter(gmc_buffer, engine="openpyxl") as writer:
                        gmc_df.to_excel(writer, sheet_name="Products", index=False)
                        if not gmc_removed.empty:
                            gmc_removed.to_excel(writer, sheet_name="Retires", index=False)
                    gmc_buffer.seek(0)   # BUG 10 corrigé
                    st.download_button(
                        label="Flux Google Merchant Center.xlsx",
                        data=gmc_buffer,
                        file_name="Flux Google Merchant Center - Products source.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
    else:
        st.info("Importez un fichier export-variants (.xlsx) pour commencer")


# ============================================================
# PAGE 2 : Apercu des fichiers actuels
# ============================================================
elif page == "Apercu des fichiers actuels":

    st.markdown('<h2 class="section-title">Etat actuel des fichiers</h2>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)

    acp_full = pd.DataFrame()
    gmc_full = pd.DataFrame()

    with col1:
        acp_full = load_acp_as_df(ACP_FEED)
        if not acp_full.empty:
            st.metric("ACP Feed", f"{len(acp_full)} produits, {len(acp_full.columns)} colonnes")
        else:
            st.warning("ACP_OpenAI_Feed.csv introuvable dans 'Files to update'")

    with col2:
        gmc_path = find_gmc_file()
        if gmc_path:
            gmc_full = load_gmc_as_df(gmc_path)
            st.metric("GMC Feed", f"{len(gmc_full)} produits, {len(gmc_full.columns)} colonnes")
        else:
            st.warning("GMC XLSX introuvable dans 'Files to update'")

    dl1, dl2 = st.columns(2)
    with dl1:
        if not acp_full.empty and os.path.exists(ACP_FEED):
            with open(ACP_FEED, "rb") as _fa:
                acp_raw_bytes = _fa.read()
            st.download_button(
                label="Telecharger ACP_OpenAI_Feed.csv",
                data=acp_raw_bytes,
                file_name="ACP_OpenAI_Feed.csv",
                mime="text/csv",
                key="dl_acp_apercu",
                use_container_width=True,
            )
    with dl2:
        if gmc_path and not gmc_full.empty:
            # Télécharger le fichier original directement (préserve les HYPERLINK headers)
            with open(gmc_path, "rb") as _f:
                gmc_raw_bytes = _f.read()
            st.download_button(
                label="Telecharger GMC.xlsx",
                data=gmc_raw_bytes,
                file_name=os.path.basename(gmc_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_gmc_apercu",
                use_container_width=True,
            )

    st.markdown("---")

    cur_tab1, cur_tab2 = st.tabs(["ACP OpenAI Feed", "Google Merchant Center"])

    with cur_tab1:
        if not acp_full.empty:
            cur_acp = acp_full.copy()   # déjà chargé via load_acp_as_df
            st.markdown(f"**{len(cur_acp)} produits, {len(cur_acp.columns)} colonnes**")
            acp_mode = st.radio("Mode", ["Consulter", "Modifier"], horizontal=True, key="acp_mode")

            if acp_mode == "Consulter":
                c_sub1, c_sub2, c_sub3, c_sub4 = st.tabs(["Tableau", "Colonnes & Remplissage", "Disponibilite", "Prix"])
                with c_sub1:
                    all_cols = list(cur_acp.columns)
                    sel = st.multiselect("Colonnes a afficher", all_cols, default=all_cols, key="cur_acp_cols")
                    search = st.text_input("Rechercher", key="cur_acp_search")
                    d = cur_acp[sel] if sel else cur_acp
                    if search:
                        d = d[d.apply(lambda r: r.astype(str).str.contains(search, case=False).any(), axis=1)]
                    st.dataframe(d, use_container_width=True, height=450)
                    st.caption(f"{len(d)} lignes")
                with c_sub2:
                    fill = column_fill_stats(cur_acp)
                    m1, m2 = st.columns(2)
                    m1.metric("Colonnes remplies", f"{(fill['Vide']=='Non').sum()}/{len(cur_acp.columns)}")
                    m2.metric("Colonnes vides", f"{(fill['Vide']=='Oui').sum()}/{len(cur_acp.columns)}")
                    st.dataframe(fill, use_container_width=True, height=450)
                with c_sub3:
                    avail_c = availability_stats(cur_acp)
                    if not avail_c.empty:
                        html_bar_chart(avail_c["Statut"].tolist(), avail_c["Nombre"].tolist())
                        st.dataframe(avail_c, use_container_width=True)
                with c_sub4:
                    ps = price_stats(cur_acp, "price")
                    if ps:
                        pcols = st.columns(len(ps))
                        for i, (k, v) in enumerate(ps.items()):
                            pcols[i].metric(k, v)

            else:  # Modifier
                st.markdown('<div style="background:#fff8e1;border:1px solid #f9a825;border-radius:6px;padding:10px 14px;margin-bottom:12px;font-size:13px;color:#6d4c00;">Modifiez les cellules directement dans le tableau. Cliquez sur <b>Sauvegarder</b> pour enregistrer.</div>', unsafe_allow_html=True)
                edit_search = st.text_input("Filtrer par SKU ou titre", key="acp_edit_search")
                edit_cols = st.multiselect(
                    "Colonnes a modifier",
                    list(cur_acp.columns),
                    default=[c for c in ["item_id","title","description","price","sale_price","availability","gtin","image_url"] if c in cur_acp.columns],
                    key="acp_edit_cols",
                )
                if edit_cols:
                    edit_df = cur_acp[edit_cols].copy()
                    if edit_search:
                        mask = edit_df.apply(lambda r: r.astype(str).str.contains(edit_search, case=False).any(), axis=1)
                        edit_idx = edit_df[mask].index
                        edit_df = edit_df.loc[edit_idx]
                    else:
                        edit_idx = edit_df.index
                    edited = st.data_editor(edit_df, use_container_width=True, height=500, num_rows="fixed", key="acp_editor")
                    b1, b2 = st.columns([1, 4])
                    with b1:
                        if st.button("Sauvegarder ACP", type="primary", key="save_acp"):
                            for col in edit_cols:
                                cur_acp.loc[edit_idx, col] = edited[col].values
                            save_acp_from_df(cur_acp, ACP_FEED)
                            st.success(f"ACP sauvegarde ({len(cur_acp)} produits)")
                            st.rerun()
                    with b2:
                        st.caption(f"{len(edit_df)} lignes affichees / {len(cur_acp)} total")
        else:
            st.info("Aucun fichier ACP trouve")

    with cur_tab2:
        gmc_path = find_gmc_file()
        if gmc_path:
            cur_gmc = load_gmc_as_df(gmc_path)   # préserve les labels HYPERLINK
            st.markdown(f"**{len(cur_gmc)} produits, {len(cur_gmc.columns)} colonnes**")
            gmc_mode = st.radio("Mode", ["Consulter", "Modifier"], horizontal=True, key="gmc_mode")

            if gmc_mode == "Consulter":
                g_sub1, g_sub2, g_sub3, g_sub4 = st.tabs(["Tableau", "Colonnes & Remplissage", "Disponibilite", "Prix"])
                with g_sub1:
                    all_g = list(cur_gmc.columns)
                    gsel = st.multiselect("Colonnes a afficher", all_g, default=all_g, key="cur_gmc_cols")
                    gsearch = st.text_input("Rechercher", key="cur_gmc_search")
                    dg = cur_gmc[gsel] if gsel else cur_gmc
                    if gsearch:
                        dg = dg[dg.apply(lambda r: r.astype(str).str.contains(gsearch, case=False).any(), axis=1)]
                    st.dataframe(dg, use_container_width=True, height=450)
                    st.caption(f"{len(dg)} lignes")
                with g_sub2:
                    gfill = column_fill_stats(cur_gmc)
                    m1, m2 = st.columns(2)
                    m1.metric("Colonnes remplies", f"{(gfill['Vide']=='Non').sum()}/{len(cur_gmc.columns)}")
                    m2.metric("Colonnes vides", f"{(gfill['Vide']=='Oui').sum()}/{len(cur_gmc.columns)}")
                    st.dataframe(gfill, use_container_width=True, height=450)
                with g_sub3:
                    gavail = availability_stats(cur_gmc)
                    if not gavail.empty:
                        html_bar_chart(gavail["Statut"].tolist(), gavail["Nombre"].tolist())
                        st.dataframe(gavail, use_container_width=True)
                with g_sub4:
                    gps = price_stats(cur_gmc, "price")
                    if gps:
                        gpcols = st.columns(len(gps))
                        for i, (k, v) in enumerate(gps.items()):
                            gpcols[i].metric(k, v)
                        prices_cur = pd.to_numeric(
                            cur_gmc["price"].str.replace(r"\s*EUR\s*$","",regex=True).str.replace(",",".").str.replace(r"[^\d.]","",regex=True),
                            errors="coerce"
                        ).dropna()
                        html_histogram(prices_cur)

            else:  # Modifier
                st.markdown('<div style="background:#fff8e1;border:1px solid #f9a825;border-radius:6px;padding:10px 14px;margin-bottom:12px;font-size:13px;color:#6d4c00;">Modifiez les cellules directement dans le tableau. Cliquez sur <b>Sauvegarder</b> pour enregistrer.</div>', unsafe_allow_html=True)
                gedit_search = st.text_input("Filtrer par SKU ou titre", key="gmc_edit_search")
                gmc_edit_defaults = ["id", "title", "description", "price", "sale price", "availability", "gtin", "image link"]
                gedit_cols = st.multiselect(
                    "Colonnes a modifier",
                    list(cur_gmc.columns),
                    default=[c for c in gmc_edit_defaults if c in cur_gmc.columns],
                    key="gmc_edit_cols",
                )
                if gedit_cols:
                    gedit_df = cur_gmc[gedit_cols].copy()
                    if gedit_search:
                        gmask = gedit_df.apply(lambda r: r.astype(str).str.contains(gedit_search, case=False).any(), axis=1)
                        gedit_idx = gedit_df[gmask].index
                        gedit_df = gedit_df.loc[gedit_idx]
                    else:
                        gedit_idx = gedit_df.index
                    gedited = st.data_editor(gedit_df, use_container_width=True, height=500, num_rows="fixed", key="gmc_editor")
                    gb1, gb2 = st.columns([1, 4])
                    with gb1:
                        if st.button("Sauvegarder GMC", type="primary", key="save_gmc"):
                            for col in gedit_cols:
                                cur_gmc.loc[gedit_idx, col] = gedited[col].values
                            save_gmc_from_df(cur_gmc, gmc_path)   # préserve les HYPERLINK headers
                            st.success(f"GMC sauvegarde ({len(cur_gmc)} produits)")
                            st.rerun()
                    with gb2:
                        st.caption(f"{len(gedit_df)} lignes affichees / {len(cur_gmc)} total")
        else:
            st.info("Aucun fichier GMC trouve")

# Footer
st.markdown('<div class="calebasse-footer">Laboratoire Calebasse &copy; 2026</div>', unsafe_allow_html=True)